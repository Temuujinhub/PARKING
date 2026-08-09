"""Тайлангийн Excel (openpyxl) рендерүүд — reports_router-ийн excel endpoint-уудын
workbook үүсгэх код. Router-т зөвхөн query/эрхийн шалгалт үлдэж, энд файл угсарна."""
import io
from datetime import datetime, timedelta

from fastapi.responses import StreamingResponse

# reports_router-тэй ижил TZ хөрвүүлэлт (өдрийн зааг локал цагаар)
from ..config import settings as _cfg
TZ = timedelta(hours=_cfg.tz_offset_hours)


def _excel_response(wb, prefix: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{prefix}_{datetime.utcnow():%Y%m%d_%H%M}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


def _xlsx(prefix, title, headers, rows, widths=None, total_row=None):
    """Нэг хуудастай Excel файл үүсгэх стандарт helper — header тод, мөрүүд, нийлбэр мөр,
    баганы өргөн. 9 Excel endpoint-ийн давхардсан boilerplate-ийг орлоно."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(list(headers))
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append(list(r))
    if total_row is not None:
        ws.append(list(total_row))
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return _excel_response(wb, prefix)


def revenue_excel(data):
    """Орлогын тайлан (зогсоолоор)."""
    def _pct(r):
        a = r.get("accrued_amount") or 0
        return round(r["paid_amount"] / a * 100, 1) if a else 0
    rows = [[r["site_name"], r["entered"], r["exited"], r["total_minutes"],
             r.get("accrued_amount", 0), r["cash_amount"],
             r["qpay_amount"], r["pos_amount"], r["transfer_amount"],
             r["paid_amount"], _pct(r), r["unpaid_amount"], r.get("debt_amount", 0)]
            for r in data["rows"]]
    t = data["totals"]
    total_row = ["НИЙТ", t["entered"], t["exited"], t["total_minutes"],
                 t.get("accrued_amount", 0), t["cash_amount"],
                 t["qpay_amount"], t["pos_amount"], t["transfer_amount"],
                 t["paid_amount"], _pct(t), t["unpaid_amount"], t.get("debt_amount", 0)]
    return _xlsx("revenue", "Орлогын тайлан",
                 ["Зогсоол", "Орсон", "Гарсан", "Нийт минут", "Үүссэн төлбөр (₮)",
                  "Бэлэн (₮)", "QPay (₮)", "Карт (₮)",
                  "Дансаар (₮)", "Нийт төлөгдсөн (₮)", "Цуглуулалт (%)",
                  "Хүлээгдэж буй (₮)", "Өр болсон (₮)"],
                 rows, widths=(30, 10, 10, 12, 18, 14, 14, 14, 14, 18, 14, 16, 16),
                 total_row=total_row)


def transactions_excel(rows):
    """Дэлгэрэнгүй бичилтийн Excel (rows = _txn_rows-ийн гаралт)."""
    headers = ["Дугаар", "Зогсоол", "Орсон", "Гарсан", "Хугацаа(мин)", "Машины төрөл",
               "Хөнгөлөлт", "Үндсэн(₮)", "Хөнгөлсөн(₮)", "НӨАТ(₮)", "Нийт(₮)", "Төлсөн(₮)",
               "Төлбөрийн хэрэгсэл", "Гүйлгээний утга", "Төлөв", "Кассчин", "ДДТД", "Сугалаа", "ТТД"]
    xrows = [[r["plate_number"], r["site_name"],
              (r["entry_time"] or "").replace("T", " ")[:16],
              (r["exit_time"] or "").replace("T", " ")[:16], r["duration_minutes"],
              r["car_type"], r["discount_name"] or "", r["base_fee"], r["discount_amount"],
              r["vat_amount"], r["total_fee"], r["paid_amount"], r["provider"] or "",
              r["invoice_no"] or "",
              r["status"], r["cashier"] or "", r["ebarimt_id"] or "", r["lottery_code"] or "",
              r["customer_tin"] or ""] for r in rows]
    return _xlsx("bichilt", "Бичилт", headers, xrows,
                 widths=(11, 14, 17, 17, 11, 12, 12, 11, 11, 9, 11, 11, 16, 26, 15, 14, 20, 12, 12))


def settlement_excel(rows):
    """Санхүүгийн тооцооны Excel (rows = settlement-ийн гаралт)."""
    xrows = [[r["date"], r["card"], r["pos_qpay"], r["qr_qpay"], r["cash"], r["transfer"],
              r["system_total"],
              r["confirmed_cash"], r["confirmed_transfer"], r["confirmed_total"],
              r["difference"], r["debt"],
              ", ".join(r["workers"]), "Хаагдсан" if r["status"] == "CLOSED" else "Нээлттэй",
              r["closed_by"] or ""] for r in rows]
    return _xlsx("monggon_tootsoo", "Мөнгөн тооцоо",
                 ["Огноо", "pos-Карт", "pos-QPay", "QR-QPay", "Бэлэн", "Дансаар", "Систем нийт",
                  "Баталгаа бэлэн", "Баталгаа данс", "Баталгаа нийт", "Зөрүү", "Өр (үүссэн)",
                  "Ажилтан", "Төлөв", "Хаасан"],
                 xrows, widths=(12, 11, 11, 11, 11, 11, 12, 13, 13, 13, 12, 12, 20, 11, 14))


def daily_excel(out, tot):
    """Өдөр өдрөөр задарсан тайлангийн Excel (out/tot = _daily_rows-ийн гаралт)."""
    rows = [[r["date"], r["entered"], r["exited"], r["cash_amount"], r["qpay_amount"],
             r["pos_amount"], r["transfer_amount"], r["paid_amount"]] for r in out]
    total_row = ["НИЙТ", tot["entered"], tot["exited"], tot["cash_amount"], tot["qpay_amount"],
                 tot["pos_amount"], tot["transfer_amount"], tot["paid_amount"]]
    return _xlsx("daily", "Өдрийн тайлан",
                 ["Огноо", "Орсон", "Гарсан", "Бэлэн (₮)", "QPay (₮)", "Карт (₮)",
                  "Дансаар (₮)", "Нийт орлого (₮)"],
                 rows, widths=(14, 10, 10, 14, 14, 14, 14, 16), total_row=total_row)


def by_shift_excel(rows):
    """Ээлжээр тайлангийн Excel (rows = _shift_rows-ийн гаралт)."""
    xrows = [[r["date"], r["window"], r["entered"], r["exited"],
              r["cash_amount"], r["qpay_amount"], r["pos_amount"], r["transfer_amount"],
              r["paid_amount"]] for r in rows]
    return _xlsx("eeljeer", "Ээлжээр",
                 ["Ээлжийн өдөр", "Зааг", "Орсон", "Гарсан", "Бэлэн (₮)", "QPay (₮)", "Карт (₮)",
                  "Дансаар (₮)", "Нийт (₮)"],
                 xrows, widths=(14, 14, 9, 9, 13, 13, 13, 13, 14))


def monthly_excel(data, daily_rows):
    """Сараар тайлангийн Excel — 2 хуудас: сарын нэгтгэл + өдрөөр задаргаа."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    # Sheet 1 — сарын нэгтгэл
    ws = wb.active
    ws.title = "Сарын нэгтгэл"
    ws.append(["Сар", "Гүйлгээ", "Бэлэн (₮)", "QPay (₮)", "Карт (₮)", "Дансаар (₮)", "Нийт орлого (₮)"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in data["rows"]:
        ws.append([r["month"], r["count"], r["cash"], r["qpay"], r["pos"], r["transfer"], r["total"]])
    t = data["totals"]
    ws.append(["НИЙТ", t["count"], t["cash"], t["qpay"], t["pos"], t["transfer"], t["total"]])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    for col, w in zip("ABCDEFG", (12, 10, 14, 14, 14, 14, 16)):
        ws.column_dimensions[col].width = w
    # Sheet 2 — доторх өдрийн задаргаа (нэгтгэлийн дэлгэрэнгүй)
    ws2 = wb.create_sheet("Өдрөөр задаргаа")
    ws2.append(["Огноо", "Орсон", "Гарсан", "Бэлэн (₮)", "QPay (₮)", "Карт (₮)",
                "Дансаар (₮)", "Нийт орлого (₮)"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    for r in daily_rows:
        ws2.append([r["date"], r["entered"], r["exited"], r["cash_amount"], r["qpay_amount"],
                    r["pos_amount"], r["transfer_amount"], r["paid_amount"]])
    for col, w in zip("ABCDEFGH", (12, 10, 10, 14, 14, 14, 14, 16)):
        ws2.column_dimensions[col].width = w
    return _excel_response(wb, "sariin_negtgel")


def by_payment_excel(data, start, end):
    """Төлбөрийн төрлөөр тайлангийн Excel (олон хэсэгтэй тул _xlsx-т багтахгүй)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook(); ws = wb.active; ws.title = "Төлбөрийн төрлөөр"
    # Хамрах хугацаа (хэднээс хэд хүртэлх өдрүүд)
    period = ws.cell(row=1, column=1,
                     value=f"Хугацаа: {start:%Y-%m-%d} – {(end - timedelta(days=1)):%Y-%m-%d}  "
                           f"({(end - start).days} хоног)")
    period.font = Font(bold=True)
    ws.append([])
    ws.append(["Төлбөрийн хэрэгсэл", "Гүйлгээ", "Дүн (₮)"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in data["by_method"]:
        ws.append([r["key"], r["count"], r["amount"]])
    ws.append([])
    ws.append(["Машины төрөл", "Тоо", "Дүн (₮)"])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    for r in data["by_car"]:
        ws.append([r["key"], r["count"], r["amount"]])
    for col, w in zip("ABC", (22, 12, 14)):
        ws.column_dimensions[col].width = w
    return _excel_response(wb, "tolboriin_torol")


def site_sessions_excel(site, sessions):
    """Нэг зогсоолын session-уудын дэлгэрэнгүй Excel."""
    STATUS_MN = {"OPEN": "Зогсож байна", "AWAITING_PAYMENT": "Төлбөр хүлээж буй",
                 "PAID": "Төлсөн", "CLOSED": "Гарсан", "FREE": "Үнэгүй гарсан",
                 "MANUAL_CLOSED": "Гараар хаасан"}
    rows = [[s.plate_number,
             (s.entry_time + TZ).strftime("%Y-%m-%d %H:%M"),
             (s.exit_time + TZ).strftime("%Y-%m-%d %H:%M") if s.exit_time else "",
             s.duration_minutes or "",
             float(s.total_fee or 0), float(s.vat_amount or 0), float(s.discount_amount or 0),
             "Тийм" if s.is_registered else "",
             STATUS_MN.get(s.status, s.status)] for s in sessions]
    return _xlsx(f"sessions_{site.site_code}", site.name[:30],
                 ["Дугаар", "Орсон", "Гарсан", "Хугацаа (мин)", "Дүн (₮)", "НӨАТ (₮)",
                  "Хөнгөлөлт (₮)", "Гэрээт", "Төлөв"],
                 rows, widths=(12, 18, 18, 14, 12, 10, 14, 8, 18))


def shifts_excel(db, shifts):
    """Касс хаалтын тайлангийн Excel (нийлбэр мөрөнд зөвхөн A, J тод тул _xlsx-гүй)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    from sqlalchemy import func as f

    from ..models import Payment
    wb = Workbook()
    ws = wb.active
    ws.title = "Касс хаалтын тайлан"
    ws.append(["Кассчин", "Төлөв", "Нээсэн цаг", "Хаасан цаг", "Эхэлсэн дүн (₮)",
               "Гүйлгээний тоо", "Бэлэн (₮)", "QPay (₮)", "Карт (₮)", "Дансаар (₮)",
               "Нийт орлого (₮)"])
    for c in ws[1]:
        c.font = Font(bold=True)
    grand = 0.0
    for s in shifts:
        totals = dict(db.query(Payment.provider, f.coalesce(f.sum(Payment.amount), 0))
                      .filter(Payment.shift_id == s.id, Payment.status == "PAID")
                      .group_by(Payment.provider).all())
        count = db.query(Payment).filter(Payment.shift_id == s.id, Payment.status == "PAID").count()
        cash, qpay_amt, pos, transfer = (float(totals.get(k, 0))
                                         for k in ("CASH", "QPAY", "POS", "TRANSFER"))
        total = cash + qpay_amt + pos + transfer
        grand += total
        ws.append([s.user.username if s.user else "", "Нээлттэй" if s.status == "OPEN" else "Хаагдсан",
                   (s.opened_at + TZ).strftime("%Y-%m-%d %H:%M"),
                   (s.closed_at + TZ).strftime("%Y-%m-%d %H:%M") if s.closed_at else "",
                   float(s.opening_amount or 0), count, cash, qpay_amt, pos, transfer, total])
    ws.append(["НИЙТ", "", "", "", "", "", "", "", "", "", grand])
    ws[f"A{ws.max_row}"].font = Font(bold=True)
    ws[f"K{ws.max_row}"].font = Font(bold=True)
    for col, w in zip("ABCDEFGHIJK", (14, 10, 18, 18, 14, 14, 12, 12, 12, 12, 16)):
        ws.column_dimensions[col].width = w
    return _excel_response(wb, "cashier_shifts")


def audit_logs_excel(logs):
    """Үйлдлийн логийн Excel."""
    import json as _json
    rows = [[(a.created_at + TZ).strftime("%Y-%m-%d %H:%M:%S"), a.username, a.action,
             a.entity or "", a.entity_id or "",
             _json.dumps(a.detail, ensure_ascii=False) if a.detail else ""] for a in logs]
    return _xlsx("uildliin_log", "Үйлдлийн лог",
                 ["Огноо", "Хэрэглэгч", "Үйлдэл", "Объект", "Объект ID", "Дэлгэрэнгүй"],
                 rows, widths=(20, 14, 18, 12, 38, 50))


def lpr_events_excel(events, names):
    """Камерын уншилтын логийн Excel."""
    rows = [[(e.created_at + TZ).strftime("%Y-%m-%d %H:%M:%S"),
             "Орох" if e.lane_dir == "entry" else "Гарах",
             e.plate_number, names.get(e.device_id, ""),
             round(e.confidence or 0), "Тийм" if e.accepted else "Үгүй",
             e.reject_reason or ""] for e in events]
    return _xlsx("kamer_unshilt", "Камерын уншилт",
                 ["Огноо", "Чиглэл", "Дугаар", "Камер", "Итгэлцүүр %", "Хүлээн авсан", "Шалтгаан"],
                 rows, widths=[20, 8, 12, 18, 12, 14, 30])


def barrier_commands_excel(cmds, plates, dev, src_mn):
    """Хаалтны командын логийн Excel."""
    rows = [[(c.created_at + TZ).strftime("%Y-%m-%d %H:%M:%S") if c.created_at else "",
             plates.get(c.session_id, ""), c.command,
             src_mn.get(c.command_source, c.command_source), c.status,
             dev.get(c.device_id, ""), c.issued_by or "", (c.response_text or "")[:120]]
            for c in cmds]
    return _xlsx("haalt_komand", "Хаалтны команд",
                 ["Огноо", "Дугаар", "Команд", "Эх сурвалж", "Төлөв", "Хаалт", "Оператор", "Хариу"],
                 rows, widths=[20, 12, 10, 22, 10, 16, 14, 30])
