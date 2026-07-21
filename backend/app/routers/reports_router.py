"""Тайлан: dashboard статистик, зогсоолын орлого, Excel экспорт, НӨАТ баримт, лог."""
import io
from datetime import datetime, timedelta

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from ..auth import require
from ..database import get_db
from ..models import (
    AuditLog, LprEvent, ParkingSession, ParkingSite, Payment, User, VatReceipt,
)
from ..serializers import to_dict

router = APIRouter(prefix="/api/reports", tags=["reports"])

# DB бүх цагийг UTC-ээр хадгалдаг; хэрэглэгч локал (УБ, UTC+8) өдрөөр сэтгэдэг тул
# өдрийн зааг, цагийн бүлэглэлтийг TZ-ээр хөрвүүлнэ.
from ..config import settings as _cfg  # noqa: E402
TZ = timedelta(hours=_cfg.tz_offset_hours)


def _local_midnight_utc(dt_utc: datetime) -> datetime:
    """Тухайн UTC моментын ЛОКАЛ өдрийн 00:00-ийг UTC-ээр буцаана."""
    return (dt_utc + TZ).replace(hour=0, minute=0, second=0, microsecond=0) - TZ


def _range(date_from: str | None, date_to: str | None):
    """UI-ийн огноог ЛОКАЛ өдөр гэж ойлгож UTC зааг руу хөрвүүлнэ."""
    start = (datetime.fromisoformat(date_from) - TZ) if date_from else _local_midnight_utc(datetime.utcnow())
    end = (datetime.fromisoformat(date_to) + timedelta(days=1) - TZ) if date_to         else datetime.utcnow() + timedelta(days=1)
    return start, end


def _daily_rows(db, start, end, site_id):
    """Өдөр өдрөөр орц/гарц + төлбөрийн хэрэгслээр (бэлэн/QPay/карт) орлого.
    daily_report ба daily_excel хоёр ижил логик ашигладаг тул нэг эх сурвалж болгов."""
    out = []
    day = _local_midnight_utc(start)
    while day < end:
        nxt = day + timedelta(days=1)
        sq = db.query(ParkingSession).filter(ParkingSession.entry_time >= day,
                                             ParkingSession.entry_time < nxt)
        pq = (db.query(Payment.provider, func.coalesce(func.sum(Payment.amount), 0))
              .join(ParkingSession, Payment.session_id == ParkingSession.id)
              .filter(Payment.status == "PAID", Payment.paid_at >= day, Payment.paid_at < nxt))
        if site_id:
            sq = sq.filter(ParkingSession.site_id == site_id)
            pq = pq.filter(ParkingSession.site_id == site_id)
        prov = dict(pq.group_by(Payment.provider).all())
        cash, qpay_amt, pos = (float(prov.get(k, 0)) for k in ("CASH", "QPAY", "POS"))
        out.append({"date": (day + TZ).strftime("%Y-%m-%d"), "entered": sq.count(),
                    "exited": sq.filter(ParkingSession.exit_time.isnot(None)).count(),
                    "cash_amount": cash, "qpay_amount": qpay_amt, "pos_amount": pos,
                    "paid_amount": cash + qpay_amt + pos})
        day = nxt
    totals = {k: sum(r[k] for r in out) for k in
              ("entered", "exited", "cash_amount", "qpay_amount", "pos_amount", "paid_amount")}
    return out, totals


def _xlsx(prefix, title, headers, rows, widths=None, total_row=None):
    """Нэг хуудастай Excel файл үүсгэх стандарт helper — header тод, мөрүүд, нийлбэр мөр,
    баганы өргөн. 9 Excel endpoint-ийн давхардсан boilerplate-ийг орлоно."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(list(headers))
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append(list(r))
    if total_row is not None:
        ws.append(list(total_row))
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return _excel_response(wb, prefix)


@router.get("/dashboard")
def dashboard_stats(db: Session = Depends(get_db), user: User = Depends(require("dashboard"))):
    """Нүүр хуудасны статистик."""
    today = _local_midnight_utc(datetime.utcnow())
    open_count = db.query(ParkingSession).filter(
        ParkingSession.status.in_(["OPEN", "AWAITING_PAYMENT", "PAID"])).count()
    awaiting = db.query(ParkingSession).filter(ParkingSession.status == "AWAITING_PAYMENT").count()
    today_entries = db.query(ParkingSession).filter(ParkingSession.entry_time >= today).count()
    today_exits = db.query(ParkingSession).filter(ParkingSession.exit_time >= today).count()
    today_revenue = float(db.query(func.coalesce(func.sum(Payment.amount), 0)).filter(
        Payment.status == "PAID", Payment.paid_at >= today).scalar())
    total_capacity = db.query(func.coalesce(func.sum(ParkingSite.capacity), 0)).filter(
        ParkingSite.is_active.is_(True)).scalar()

    sites = []
    for s in db.query(ParkingSite).filter(ParkingSite.is_active.is_(True)).all():
        occupied = db.query(ParkingSession).filter(
            ParkingSession.site_id == s.id,
            ParkingSession.status.in_(["OPEN", "AWAITING_PAYMENT", "PAID"])).count()
        revenue = float(db.query(func.coalesce(func.sum(Payment.amount), 0))
                        .join(ParkingSession, Payment.session_id == ParkingSession.id)
                        .filter(ParkingSession.site_id == s.id, Payment.status == "PAID",
                                Payment.paid_at >= today).scalar())
        sites.append({"id": s.id, "name": s.name, "capacity": s.capacity,
                      # capacity=0 → дүүргэлтгүй: сул тоо null
                      "occupied": occupied,
                      "free": max(0, s.capacity - occupied) if s.capacity else None,
                      "today_revenue": revenue})

    # Сүүлийн 7 хоногийн орлого (график)
    week = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        rev = float(db.query(func.coalesce(func.sum(Payment.amount), 0)).filter(
            Payment.status == "PAID", Payment.paid_at >= day,
            Payment.paid_at < day + timedelta(days=1)).scalar())
        week.append({"date": (day + TZ).strftime("%m-%d"), "revenue": rev})

    # Өнөөдрийн цагийн ачаалал — цаг тус бүрийн орц/гарц (0–23)
    from sqlalchemy import Integer, cast
    hourly = {h: {"hour": h, "entries": 0, "exits": 0} for h in range(24)}
    for hr, cnt in (db.query(cast(func.extract("hour", ParkingSession.entry_time + TZ), Integer),
                             func.count()).filter(ParkingSession.entry_time >= today)
                    .group_by(func.extract("hour", ParkingSession.entry_time + TZ)).all()):
        if hr is not None:
            hourly[int(hr)]["entries"] = int(cnt)
    for hr, cnt in (db.query(cast(func.extract("hour", ParkingSession.exit_time + TZ), Integer),
                             func.count()).filter(ParkingSession.exit_time >= today)
                    .group_by(func.extract("hour", ParkingSession.exit_time + TZ)).all()):
        if hr is not None:
            hourly[int(hr)]["exits"] = int(cnt)
    hourly_load = [hourly[h] for h in range(24)]

    # Төхөөрөмжийн холболтын статус (сүүлийн 3 минутад холбогдсон = онлайн)
    from ..models import Device
    online_cutoff = datetime.utcnow() - timedelta(minutes=3)
    devices = db.query(Device).filter(Device.status == "active").all()
    device_status = []
    online_n = 0
    for d in devices:
        online = bool(d.last_seen and d.last_seen >= online_cutoff)
        if online:
            online_n += 1
        device_status.append({
            "id": d.id, "name": d.name, "device_type": d.device_type,
            "lane_dir": d.lane_dir, "site_name": d.site.name if d.site else None,
            "online": online, "last_seen": d.last_seen.isoformat() if d.last_seen else None,
        })

    # Ажиллаж буй ээлж — хэн аль зогсоолд POS/системд нэвтэрч ажиллаж байгаа
    from ..models import CashierShift
    active_shifts = []
    for sh in (db.query(CashierShift).filter(CashierShift.status == "OPEN")
               .order_by(CashierShift.opened_at.desc()).all()):
        rev = float(db.query(func.coalesce(func.sum(Payment.amount), 0)).filter(
            Payment.status == "PAID", Payment.cashier_id == sh.user_id,
            Payment.paid_at >= sh.opened_at).scalar() or 0)
        active_shifts.append({
            "cashier": (sh.user.full_name or sh.user.username) if sh.user else "?",
            "site_name": sh.site.name if sh.site else "Бүх зогсоол",
            "opened_at": sh.opened_at.isoformat(), "revenue": rev})

    # Төхөөрөмжийн төрлөөр (карт дээр том тоогоор харуулна)
    cameras_total = sum(1 for d in device_status if d["device_type"] == "camera")
    barriers_total = sum(1 for d in device_status if d["device_type"] == "barrier")
    return {"open_sessions": open_count, "awaiting_payment": awaiting,
            "today_entries": today_entries, "today_exits": today_exits,
            "today_revenue": today_revenue, "total_capacity": int(total_capacity or 0),
            "sites": sites, "week_revenue": week, "hourly_load": hourly_load,
            "sites_total": len(sites), "active_shifts": active_shifts,
            "cameras_total": cameras_total, "barriers_total": barriers_total,
            "devices_online": online_n, "devices_total": len(devices),
            "device_status": device_status}


@router.get("/revenue")
def revenue_report(date_from: str | None = None, date_to: str | None = None,
                   site_id: str | None = None,
                   db: Session = Depends(get_db), user: User = Depends(require("reports"))):
    """Зогсоол тус бүрийн орлогын тайлан (easy-park 'Зогсоолын төлбөрийн тайлан')."""
    start, end = _range(date_from, date_to)
    out = []
    sites = db.query(ParkingSite).all()
    if site_id:
        sites = [s for s in sites if s.id == site_id]
    for s in sites:
        base = db.query(ParkingSession).filter(ParkingSession.site_id == s.id,
                                               ParkingSession.entry_time >= start,
                                               ParkingSession.entry_time < end)
        entered = base.count()
        exited = base.filter(ParkingSession.exit_time.isnot(None)).count()
        minutes = db.query(func.coalesce(func.sum(ParkingSession.duration_minutes), 0)).filter(
            ParkingSession.site_id == s.id, ParkingSession.entry_time >= start,
            ParkingSession.entry_time < end).scalar()
        # Төлбөрийн төрлөөр задаргаа (easy-park UAT items 1, 4, 6, 7)
        prov = dict(db.query(Payment.provider, func.coalesce(func.sum(Payment.amount), 0))
                    .join(ParkingSession, Payment.session_id == ParkingSession.id)
                    .filter(ParkingSession.site_id == s.id, Payment.status == "PAID",
                            Payment.paid_at >= start, Payment.paid_at < end)
                    .group_by(Payment.provider).all())
        cash, qpay_amt, pos = (float(prov.get(k, 0)) for k in ("CASH", "QPAY", "POS"))
        paid = cash + qpay_amt + pos
        unpaid = float(db.query(func.coalesce(func.sum(ParkingSession.total_fee), 0)).filter(
            ParkingSession.site_id == s.id, ParkingSession.status == "AWAITING_PAYMENT",
            ParkingSession.entry_time >= start, ParkingSession.entry_time < end).scalar())
        out.append({"site_id": s.id, "site_name": s.name, "entered": entered, "exited": exited,
                    "total_minutes": int(minutes or 0),
                    "cash_amount": cash, "qpay_amount": qpay_amt, "pos_amount": pos,
                    "paid_amount": paid, "unpaid_amount": unpaid})
    totals = {
        "entered": sum(r["entered"] for r in out), "exited": sum(r["exited"] for r in out),
        "total_minutes": sum(r["total_minutes"] for r in out),
        "cash_amount": sum(r["cash_amount"] for r in out),
        "qpay_amount": sum(r["qpay_amount"] for r in out),
        "pos_amount": sum(r["pos_amount"] for r in out),
        "paid_amount": sum(r["paid_amount"] for r in out),
        "unpaid_amount": sum(r["unpaid_amount"] for r in out),
    }
    return {"rows": out, "totals": totals,
            "date_from": start.isoformat(), "date_to": end.isoformat()}


@router.get("/revenue/excel")
def revenue_excel(date_from: str | None = None, date_to: str | None = None,
                  db: Session = Depends(get_db), user: User = Depends(require("reports"))):
    data = revenue_report(date_from, date_to, None, db, user)
    rows = [[r["site_name"], r["entered"], r["exited"], r["total_minutes"], r["cash_amount"],
             r["qpay_amount"], r["pos_amount"], r["paid_amount"], r["unpaid_amount"]]
            for r in data["rows"]]
    t = data["totals"]
    total_row = ["НИЙТ", t["entered"], t["exited"], t["total_minutes"], t["cash_amount"],
                 t["qpay_amount"], t["pos_amount"], t["paid_amount"], t["unpaid_amount"]]
    return _xlsx("revenue", "Орлогын тайлан",
                 ["Зогсоол", "Орсон", "Гарсан", "Нийт минут", "Бэлэн (₮)", "QPay (₮)", "Карт (₮)",
                  "Нийт төлөгдсөн (₮)", "Төлөгдөөгүй (₮)"],
                 rows, widths=(30, 10, 10, 12, 14, 14, 14, 18, 16), total_row=total_row)


@router.get("/daily")
def daily_report(date_from: str | None = None, date_to: str | None = None,
                 site_id: str | None = None,
                 db: Session = Depends(get_db), user: User = Depends(require("reports"))):
    """Өдөр өдрөөр задарсан тайлан (easy-park UAT item 3)."""
    start, end = _range(date_from, date_to)
    out, totals = _daily_rows(db, start, end, site_id)
    return {"rows": out, "totals": totals}


@router.get("/monthly")
def monthly_report(date_from: str | None = None, date_to: str | None = None,
                   site_id: str | None = None,
                   db: Session = Depends(get_db), user: User = Depends(require("reports"))):
    """Сар сараар — төлбөрийн хэрэгслээр (бэлэн/QPay/карт) задарсан тайлан."""
    from sqlalchemy import Integer, cast
    start, end = _range(date_from, date_to)
    ymexpr = (cast(func.extract("year", Payment.paid_at + TZ), Integer) * 100
              + cast(func.extract("month", Payment.paid_at + TZ), Integer))
    q = (db.query(ymexpr.label("ym"), Payment.provider,
                  func.coalesce(func.sum(Payment.amount), 0), func.count())
         .join(ParkingSession, Payment.session_id == ParkingSession.id)
         .filter(Payment.status == "PAID", Payment.paid_at >= start, Payment.paid_at < end))
    if site_id:
        q = q.filter(ParkingSession.site_id == site_id)
    months = {}
    for ym, prov, amt, cnt in q.group_by("ym", Payment.provider).all():
        m = months.setdefault(int(ym), {"cash": 0.0, "qpay": 0.0, "pos": 0.0, "count": 0})
        key = {"CASH": "cash", "QPAY": "qpay", "POS": "pos"}.get(prov)
        if key:
            m[key] += float(amt)
        m["count"] += int(cnt)
    out = []
    for ym in sorted(months, reverse=True):
        m = months[ym]
        out.append({"month": f"{ym // 100}-{ym % 100:02d}", **m,
                    "total": m["cash"] + m["qpay"] + m["pos"]})
    totals = {k: sum(r[k] for r in out) for k in ("cash", "qpay", "pos", "total", "count")}
    return {"rows": out, "totals": totals}


PROVIDER_MN = {"CASH": "Бэлэн", "QPAY": "QPay", "POS": "Банкны карт"}
STATUS_MN2 = {"PAID": "Төлсөн", "FREE": "Үнэгүй", "AWAITING_PAYMENT": "Төлбөр хүлээж буй",
              "OPEN": "Нээлттэй", "CLOSED": "Хаагдсан"}


def _car_type(s) -> str:
    if s.is_registered:
        return "Гэрээт"
    if s.discount_id:
        return "Хөнгөлөлттэй"
    return "Энгийн"


def _txn_query(db, start, end, site_id, provider, car_type, status, date_field="entry"):
    """Бичилтийн шүүлттэй session query. date_field='entry'=орсон цагаар,
    'paid'=төлбөрийн огноогоор (тухайн өдрийн ГҮЙЛГЭЭ — орлоготой таарна)."""
    if date_field == "paid":
        # Тухайн хугацаанд ТӨЛӨГДСӨН гүйлгээтэй session-ууд
        paid_sub = (db.query(Payment.session_id).filter(
            Payment.status == "PAID", Payment.paid_at >= start, Payment.paid_at < end).subquery())
        q = db.query(ParkingSession).filter(ParkingSession.id.in_(db.query(paid_sub.c.session_id)))
    else:
        q = db.query(ParkingSession).filter(ParkingSession.entry_time >= start,
                                            ParkingSession.entry_time < end)
    if site_id:
        q = q.filter(ParkingSession.site_id == site_id)
    if status:
        q = q.filter(ParkingSession.status == status)
    if car_type == "contract":
        q = q.filter(ParkingSession.is_registered.is_(True))
    elif car_type == "discount":
        q = q.filter(ParkingSession.discount_id.isnot(None))
    elif car_type == "normal":
        q = q.filter(ParkingSession.is_registered.is_(False), ParkingSession.discount_id.is_(None))
    if provider:
        sub = (db.query(Payment.session_id).filter(Payment.status == "PAID",
                                                   Payment.provider == provider).subquery())
        q = q.filter(ParkingSession.id.in_(db.query(sub.c.session_id)))
    return q  # order-гүй — caller шаардлагатай бол .order_by нэмнэ


def _txn_rows(db, sessions):
    """Session жагсаалтыг бүрэн бичилт болгон дэлгэнэ (payment/receipt/cashier багцаар)."""
    from ..models import CashierShift, User, VatReceipt
    ids = [s.id for s in sessions]
    pays_by_sess = {}
    if ids:
        for p in db.query(Payment).filter(Payment.session_id.in_(ids)).all():
            pays_by_sess.setdefault(p.session_id, []).append(p)
    rec_by_sess = {r.session_id: r for r in
                   db.query(VatReceipt).filter(VatReceipt.session_id.in_(ids)).all()} if ids else {}
    cashier_ids = {p.cashier_id for ps in pays_by_sess.values() for p in ps if p.cashier_id}
    cashiers = {u.id: u.full_name or u.username for u in
                db.query(User).filter(User.id.in_(cashier_ids)).all()} if cashier_ids else {}
    out = []
    for s in sessions:
        pays = pays_by_sess.get(s.id, [])
        paid = [p for p in pays if p.status == "PAID"]
        primary = (paid[0] if paid else (pays[0] if pays else None))
        paid_amount = sum(float(p.amount) for p in paid)
        rec = rec_by_sess.get(s.id)
        out.append({
            "session_id": s.id,
            "plate_number": s.plate_number,
            "site_name": s.site.name if s.site else None,
            "entry_time": s.entry_time.isoformat() if s.entry_time else None,
            "exit_time": s.exit_time.isoformat() if s.exit_time else None,
            "duration_minutes": s.duration_minutes,
            "car_type": _car_type(s),
            "discount_name": s.discount.name if s.discount else None,
            "base_fee": float(s.base_fee or 0),
            "discount_amount": float(s.discount_amount or 0),
            "vat_amount": float(s.vat_amount or 0),
            "total_fee": float(s.total_fee or 0),
            "paid_amount": paid_amount,
            "provider": PROVIDER_MN.get(primary.provider, primary.provider) if primary else None,
            "payment_method": primary.payment_method if primary else None,
            "status": STATUS_MN2.get(s.status, s.status),
            "cashier": cashiers.get(primary.cashier_id) if primary and primary.cashier_id else None,
            "ebarimt_id": rec.ebarimt_id if rec else None,
            "lottery_code": rec.lottery_code if rec else None,
            "customer_tin": rec.customer_tin if rec else (primary.customer_tin if primary else None),
            "paid_at": primary.paid_at.isoformat() if primary and primary.paid_at else None,
        })
    return out


@router.get("/transactions")
def transactions(date_from: str | None = None, date_to: str | None = None,
                 site_id: str | None = None, provider: str | None = None,
                 car_type: str | None = None, status: str | None = None,
                 date_field: str = "entry", limit: int = 500, offset: int = 0,
                 db: Session = Depends(get_db), user: User = Depends(require("reports"))):
    """Дэлгэрэнгүй бичилтийн тайлан — машин бүрийн бүрэн мөчлөг, олон талбараар шүүнэ.
    Шүүлт: огноо (date_field=entry орсон / paid төлсөн), зогсоол, төлбөрийн хэрэгсэл,
    машины төрөл (contract/discount/normal), төлөв. Багцалж татахад /transactions/excel."""
    start, end = _range(date_from, date_to)
    q = _txn_query(db, start, end, site_id, provider, car_type, status, date_field)
    total = q.count()
    paid_sum = float(q.with_entities(func.coalesce(func.sum(ParkingSession.total_fee), 0)).scalar() or 0)
    sessions = q.order_by(ParkingSession.entry_time.desc()).offset(offset).limit(min(limit, 2000)).all()
    rows = _txn_rows(db, sessions)
    return {"total": total, "rows": rows,
            "totals": {"count": total, "total_fee": paid_sum}}


@router.get("/transactions/excel")
def transactions_excel(date_from: str | None = None, date_to: str | None = None,
                       site_id: str | None = None, provider: str | None = None,
                       car_type: str | None = None, status: str | None = None,
                       date_field: str = "entry",
                       db: Session = Depends(get_db), user: User = Depends(require("reports"))):
    """Шүүсэн бичилтүүдийг Excel болгон багцалж татна (одоогийн шүүлтээр)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    start, end = _range(date_from, date_to)
    sessions = (_txn_query(db, start, end, site_id, provider, car_type, status, date_field)
                .order_by(ParkingSession.entry_time.desc()).limit(20000).all())
    rows = _txn_rows(db, sessions)
    wb = Workbook()
    ws = wb.active
    ws.title = "Бичилт"
    headers = ["Дугаар", "Зогсоол", "Орсон", "Гарсан", "Хугацаа(мин)", "Машины төрөл",
               "Хөнгөлөлт", "Үндсэн(₮)", "Хөнгөлсөн(₮)", "НӨАТ(₮)", "Нийт(₮)", "Төлсөн(₮)",
               "Төлбөрийн хэрэгсэл", "Төлөв", "Кассчин", "ДДТД", "Сугалаа", "ТТД"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r["plate_number"], r["site_name"],
                   (r["entry_time"] or "").replace("T", " ")[:16],
                   (r["exit_time"] or "").replace("T", " ")[:16], r["duration_minutes"],
                   r["car_type"], r["discount_name"] or "", r["base_fee"], r["discount_amount"],
                   r["vat_amount"], r["total_fee"], r["paid_amount"], r["provider"] or "",
                   r["status"], r["cashier"] or "", r["ebarimt_id"] or "", r["lottery_code"] or "",
                   r["customer_tin"] or ""])
    for col, w in zip("ABCDEFGHIJKLMNOPQR",
                      (11, 14, 17, 17, 11, 12, 12, 11, 11, 9, 11, 11, 16, 15, 14, 20, 12, 12)):
        ws.column_dimensions[col].width = w
    return _excel_response(wb, "bichilt")


@router.get("/by-payment")
def by_payment(date_from: str | None = None, date_to: str | None = None, site_id: str | None = None,
               db: Session = Depends(get_db), user: User = Depends(require("reports"))):
    """Төлбөрийн төрлөөр — бүгд ТӨЛӨГДСӨН гүйлгээгээр (paid_at), тул хэрэгслээр ба
    машины төрлөөр 2 задаргаа ИЖИЛ нийлбэрт нийлнэ (тэнцвэржинэ).
    Үнэгүй гарсан нь орлогогүй тул тусад нь тоогоор (info) харуулна."""
    start, end = _range(date_from, date_to)
    # Хэрэгслээр — төлсөн гүйлгээ
    pq = (db.query(Payment.provider, func.coalesce(func.sum(Payment.amount), 0), func.count())
          .join(ParkingSession, Payment.session_id == ParkingSession.id)
          .filter(Payment.status == "PAID", Payment.paid_at >= start, Payment.paid_at < end))
    if site_id:
        pq = pq.filter(ParkingSession.site_id == site_id)
    by_method = [{"key": PROVIDER_MN.get(p, p), "amount": float(a), "count": int(c)}
                 for p, a, c in pq.group_by(Payment.provider).all()]
    # Машины төрлөөр — ИЖИЛ төлсөн гүйлгээг session-ий төрлөөр бүлэглэнэ (тэнцвэржинэ)
    payq = (db.query(Payment).join(ParkingSession, Payment.session_id == ParkingSession.id)
            .filter(Payment.status == "PAID", Payment.paid_at >= start, Payment.paid_at < end))
    if site_id:
        payq = payq.filter(ParkingSession.site_id == site_id)
    buckets = {"Гэрээт": [0, 0.0], "Хөнгөлөлттэй": [0, 0.0], "Энгийн": [0, 0.0]}
    for p in payq.all():
        buckets[_car_type(p.session)][0] += 1
        buckets[_car_type(p.session)][1] += float(p.amount)
    by_car = [{"key": k, "count": v[0], "amount": v[1]} for k, v in buckets.items()]
    # Үнэгүй гарсан машин (орлогогүй — тусад нь тоо) — гарсан огноогоор
    free_q = db.query(ParkingSession).filter(ParkingSession.status == "FREE",
                                             ParkingSession.exit_time >= start,
                                             ParkingSession.exit_time < end)
    if site_id:
        free_q = free_q.filter(ParkingSession.site_id == site_id)
    total = sum(m["amount"] for m in by_method)
    return {"by_method": by_method, "by_car": by_car,
            "total": total, "free_count": free_q.count()}


def _shift_rows(db, start, end, site_id):
    """Ээлжийн өдрөөр (өдрийг shift_change_hour-аар тасалж) төлбөрийг задлана.
    Ээлжийн өдөр D = [D + Hц, D+1 + Hц). Өдөрөөртэй ижил бүтэц, зөвхөн зааг цаг өөр."""
    from ..config import settings
    h = settings.shift_change_hour
    out = []
    # Ээлж солигдох цаг H нь ЛОКАЛ цаг — локал өдрийн эхлэл дээр H нэмж UTC зааг гаргана
    day = _local_midnight_utc(start) + timedelta(hours=h)
    if start < day:
        day -= timedelta(days=1)
    while day < end:
        nxt = day + timedelta(days=1)
        sq = db.query(ParkingSession).filter(ParkingSession.entry_time >= day,
                                             ParkingSession.entry_time < nxt)
        pq = (db.query(Payment.provider, func.coalesce(func.sum(Payment.amount), 0))
              .join(ParkingSession, Payment.session_id == ParkingSession.id)
              .filter(Payment.status == "PAID", Payment.paid_at >= day, Payment.paid_at < nxt))
        if site_id:
            sq = sq.filter(ParkingSession.site_id == site_id)
            pq = pq.filter(ParkingSession.site_id == site_id)
        prov = dict(pq.group_by(Payment.provider).all())
        cash, qpay_amt, pos = (float(prov.get(k, 0)) for k in ("CASH", "QPAY", "POS"))
        out.append({"date": (day + TZ).strftime("%Y-%m-%d"),
                    "window": f"{h:02d}:00–{h:02d}:00",
                    "entered": sq.count(),
                    "exited": sq.filter(ParkingSession.exit_time.isnot(None)).count(),
                    "cash_amount": cash, "qpay_amount": qpay_amt, "pos_amount": pos,
                    "paid_amount": cash + qpay_amt + pos})
        day = nxt
    return out


@router.get("/by-shift")
def by_shift_report(date_from: str | None = None, date_to: str | None = None,
                    site_id: str | None = None,
                    db: Session = Depends(get_db), user: User = Depends(require("reports"))):
    """Ээлжээр — өдрийг ээлж солигдох цагаар тасалж бүлэглэсэн орлого (Өдрөөртэй адил
    боловч зааг нь шөнө дунд биш, ээлж солигдох цаг)."""
    from ..config import settings
    start, end = _range(date_from, date_to)
    out = _shift_rows(db, start, end, site_id)
    totals = {k: sum(r[k] for r in out) for k in
              ("entered", "exited", "cash_amount", "qpay_amount", "pos_amount", "paid_amount")}
    return {"rows": out, "shift_hour": settings.shift_change_hour, "totals": totals}


@router.get("/settlement")
def settlement(site_id: str, date_from: str | None = None, date_to: str | None = None,
               db: Session = Depends(get_db), user: User = Depends(require("reports"))):
    """Санхүүгийн өдрийн тооцоо — pos-Карт / pos-QPay / QR-QPay / Бэлэн задаргаатай.
    Карт ба QPay нь электрон баталгаажсан тул систем=баталгаа (засахгүй); зөвхөн бэлэнг
    санхүү дансны хуулгаас баталгаажуулна. Мөн ажилтан + тухайн өдөр үүссэн өрийн дүн."""
    from ..models import CashierShift, Compensation, DailySettlement
    start, end = _range(date_from, date_to)
    setts = {s.date: s for s in db.query(DailySettlement).filter(
        DailySettlement.site_id == site_id).all()}
    out = []
    day = _local_midnight_utc(start)
    while day < end:
        nxt = day + timedelta(days=1)
        ds = (day + TZ).strftime("%Y-%m-%d")
        # Төлбөрийг provider+source-оор
        rows = (db.query(Payment.provider, Payment.source, func.coalesce(func.sum(Payment.amount), 0))
                .join(ParkingSession, Payment.session_id == ParkingSession.id)
                .filter(Payment.status == "PAID", Payment.paid_at >= day, Payment.paid_at < nxt,
                        ParkingSession.site_id == site_id)
                .group_by(Payment.provider, Payment.source).all())
        card = pos_qpay = qr_qpay = cash = 0.0
        for prov, src, amt in rows:
            amt = float(amt)
            if prov == "POS":
                card += amt
            elif prov == "QPAY":
                if src == "POS":
                    pos_qpay += amt
                else:
                    qr_qpay += amt
            elif prov == "CASH":
                cash += amt
        st_total = card + pos_qpay + qr_qpay + cash
        # Тухайн өдөр үүссэн өр (нөхөн төлбөр)
        debt = float(db.query(func.coalesce(func.sum(Compensation.amount), 0)).filter(
            Compensation.site_id == site_id, Compensation.created_at >= day,
            Compensation.created_at < nxt).scalar() or 0)
        # Ажилласан ажилтнууд (тухайн өдөр ээлж нээсэн)
        workers = sorted({(sh.user.full_name or sh.user.username) for sh in db.query(CashierShift).filter(
            CashierShift.site_id == site_id, CashierShift.opened_at >= day,
            CashierShift.opened_at < nxt).all() if sh.user})
        st = setts.get(ds)
        if st_total <= 0 and debt <= 0 and not st:
            day = nxt
            continue
        confirmed_cash = float(st.confirmed_cash) if st else 0.0
        # Карт/QPay электрон баталгаажсан = систем; зөвхөн бэлэн санхүү баталгаажуулна
        confirmed_total = card + pos_qpay + qr_qpay + confirmed_cash
        out.append({"date": ds, "card": card, "pos_qpay": pos_qpay, "qr_qpay": qr_qpay,
                    "cash": cash, "system_total": st_total, "confirmed_cash": confirmed_cash,
                    "confirmed_total": confirmed_total, "difference": st_total - confirmed_total,
                    "debt": debt, "workers": workers,
                    "status": st.status if st else "OPEN", "note": (st.note if st else "") or "",
                    "closed_by": st.closed_by if st else None,
                    "closed_at": st.closed_at.isoformat() if st and st.closed_at else None})
        day = nxt
    out.reverse()
    return {"rows": out}


@router.put("/settlement")
def settlement_upsert(body: dict, db: Session = Depends(get_db), user: User = Depends(require("reports"))):
    """Санхүү тухайн өдрийн баталгаажсан дүнг оруулж/тооцоо хаана.
    body: {site_id, date, confirmed_card?, confirmed_qpay?, confirmed_cash?, note?, status?}."""
    from ..models import DailySettlement
    if not body.get("site_id") or not body.get("date"):
        raise HTTPException(400, "site_id болон date шаардлагатай")
    st = (db.query(DailySettlement)
          .filter(DailySettlement.site_id == body["site_id"], DailySettlement.date == body["date"]).first())
    if not st:
        st = DailySettlement(site_id=body["site_id"], date=body["date"])
        db.add(st)
    for k in ("confirmed_card", "confirmed_qpay", "confirmed_cash"):
        if k in body:
            setattr(st, k, body[k] or 0)
    if "note" in body:
        st.note = body["note"]
    if body.get("status") == "CLOSED" and st.status != "CLOSED":
        st.status, st.closed_by, st.closed_at = "CLOSED", user.username, datetime.utcnow()
    elif body.get("status") == "OPEN":
        st.status, st.closed_by, st.closed_at = "OPEN", None, None
    db.add(AuditLog(username=user.username, action="SETTLEMENT", entity="settlement",
                    entity_id=f"{body['site_id']}:{body['date']}",
                    detail={"status": st.status}))
    db.commit()
    return to_dict(st)


@router.get("/settlement/excel")
def settlement_excel(site_id: str, date_from: str | None = None, date_to: str | None = None,
                     db: Session = Depends(get_db), user: User = Depends(require("reports"))):
    """Санхүүгийн тооцооны Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    data = settlement(site_id, date_from, date_to, db, user)
    wb = Workbook(); ws = wb.active; ws.title = "Мөнгөн тооцоо"
    ws.append(["Огноо", "pos-Карт", "pos-QPay", "QR-QPay", "Бэлэн", "Систем нийт",
               "Баталгаа бэлэн", "Баталгаа нийт", "Зөрүү", "Өр (үүссэн)", "Ажилтан", "Төлөв", "Хаасан"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in data["rows"]:
        ws.append([r["date"], r["card"], r["pos_qpay"], r["qr_qpay"], r["cash"], r["system_total"],
                   r["confirmed_cash"], r["confirmed_total"], r["difference"], r["debt"],
                   ", ".join(r["workers"]), "Хаагдсан" if r["status"] == "CLOSED" else "Нээлттэй",
                   r["closed_by"] or ""])
    for col, w in zip("ABCDEFGHIJKLM", (12, 11, 11, 11, 11, 12, 13, 13, 12, 12, 20, 11, 14)):
        ws.column_dimensions[col].width = w
    return _excel_response(wb, "monggon_tootsoo")


def _excel_response(wb, prefix: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{prefix}_{datetime.utcnow():%Y%m%d_%H%M}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.get("/daily/excel")
def daily_excel(date_from: str | None = None, date_to: str | None = None,
                site_id: str | None = None,
                db: Session = Depends(get_db), user: User = Depends(require("reports"))):
    """Өдөр өдрөөр задарсан тайлангийн Excel."""
    start, end = _range(date_from, date_to)
    out, tot = _daily_rows(db, start, end, site_id)
    rows = [[r["date"], r["entered"], r["exited"], r["cash_amount"], r["qpay_amount"],
             r["pos_amount"], r["paid_amount"]] for r in out]
    total_row = ["НИЙТ", tot["entered"], tot["exited"], tot["cash_amount"], tot["qpay_amount"],
                 tot["pos_amount"], tot["paid_amount"]]
    return _xlsx("daily", "Өдрийн тайлан",
                 ["Огноо", "Орсон", "Гарсан", "Бэлэн (₮)", "QPay (₮)", "Карт (₮)", "Нийт орлого (₮)"],
                 rows, widths=(14, 10, 10, 14, 14, 14, 16), total_row=total_row)


@router.get("/by-shift/excel")
def by_shift_excel(date_from: str | None = None, date_to: str | None = None, site_id: str | None = None,
                   db: Session = Depends(get_db), user: User = Depends(require("reports"))):
    """Ээлжээр тайлангийн Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    start, end = _range(date_from, date_to)
    rows = _shift_rows(db, start, end, site_id)
    wb = Workbook(); ws = wb.active; ws.title = "Ээлжээр"
    ws.append(["Ээлжийн өдөр", "Зааг", "Орсон", "Гарсан", "Бэлэн (₮)", "QPay (₮)", "Карт (₮)", "Нийт (₮)"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r["date"], r["window"], r["entered"], r["exited"],
                   r["cash_amount"], r["qpay_amount"], r["pos_amount"], r["paid_amount"]])
    for col, w in zip("ABCDEFGH", (14, 14, 9, 9, 13, 13, 13, 14)):
        ws.column_dimensions[col].width = w
    return _excel_response(wb, "eeljeer")


@router.get("/monthly/excel")
def monthly_excel(date_from: str | None = None, date_to: str | None = None, site_id: str | None = None,
                  db: Session = Depends(get_db), user: User = Depends(require("reports"))):
    """Сараар тайлангийн Excel (төлбөрийн хэрэгслээр)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    start, end = _range(date_from, date_to)
    data = monthly_report(date_from, date_to, site_id, db, user)
    wb = Workbook()
    # Sheet 1 — сарын нэгтгэл
    ws = wb.active
    ws.title = "Сарын нэгтгэл"
    ws.append(["Сар", "Гүйлгээ", "Бэлэн (₮)", "QPay (₮)", "Карт (₮)", "Нийт орлого (₮)"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in data["rows"]:
        ws.append([r["month"], r["count"], r["cash"], r["qpay"], r["pos"], r["total"]])
    t = data["totals"]
    ws.append(["НИЙТ", t["count"], t["cash"], t["qpay"], t["pos"], t["total"]])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    for col, w in zip("ABCDEF", (12, 10, 14, 14, 14, 16)):
        ws.column_dimensions[col].width = w
    # Sheet 2 — доторх өдрийн задаргаа (нэгтгэлийн дэлгэрэнгүй)
    ws2 = wb.create_sheet("Өдрөөр задаргаа")
    ws2.append(["Огноо", "Орсон", "Гарсан", "Бэлэн (₮)", "QPay (₮)", "Карт (₮)", "Нийт орлого (₮)"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    out2, _ = _daily_rows(db, start, end, site_id)
    for r in out2:
        ws2.append([r["date"], r["entered"], r["exited"], r["cash_amount"], r["qpay_amount"],
                    r["pos_amount"], r["paid_amount"]])
    for col, w in zip("ABCDEFG", (12, 10, 10, 14, 14, 14, 16)):
        ws2.column_dimensions[col].width = w
    return _excel_response(wb, "sariin_negtgel")


@router.get("/by-payment/excel")
def by_payment_excel(date_from: str | None = None, date_to: str | None = None, site_id: str | None = None,
                     db: Session = Depends(get_db), user: User = Depends(require("reports"))):
    """Төлбөрийн төрлөөр тайлангийн Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    data = by_payment(date_from, date_to, site_id, db, user)
    start, end = _range(date_from, date_to)
    wb = Workbook(); ws = wb.active; ws.title = "Төлбөрийн төрлөөр"
    # Хамрах хугацаа (хэднээс хэд хүртэлх өдрүүд)
    period = ws.cell(row=1, column=1,
                     value=f"Хугацаа: {start:%Y-%m-%d} – {(end - timedelta(days=1)):%Y-%m-%d}  "
                           f"({(end - start).days} хоног)")
    period.font = Font(bold=True)
    ws.append([])
    ws.append(["Төлбөрийн хэрэгсэл", "Гүйлгээ", "Дүн (₮)"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in data["by_method"]:
        ws.append([r["key"], r["count"], r["amount"]])
    ws.append([])
    ws.append(["Машины төрөл", "Тоо", "Дүн (₮)"])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    for r in data["by_car"]:
        ws.append([r["key"], r["count"], r["amount"]])
    for col, w in zip("ABC", (22, 12, 14)):
        ws.column_dimensions[col].width = w
    return _excel_response(wb, "tolboriin_torol")


@router.get("/site-sessions/excel")
def site_sessions_excel(site_id: str, date_from: str | None = None, date_to: str | None = None,
                        db: Session = Depends(get_db), user: User = Depends(require("reports"))):
    """Нэг зогсоолын session-уудын дэлгэрэнгүй Excel (тайлангийн мөрийн 'Татах' үйлдэл)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    start, end = _range(date_from, date_to)
    site = db.get(ParkingSite, site_id)
    if not site:
        raise HTTPException(404, "Зогсоол олдсонгүй")
    rows = (db.query(ParkingSession)
            .filter(ParkingSession.site_id == site_id,
                    ParkingSession.entry_time >= start, ParkingSession.entry_time < end)
            .order_by(ParkingSession.entry_time.desc()).limit(20000).all())
    STATUS_MN = {"OPEN": "Зогсож байна", "AWAITING_PAYMENT": "Төлбөр хүлээж буй",
                 "PAID": "Төлсөн", "CLOSED": "Гарсан", "FREE": "Үнэгүй гарсан",
                 "MANUAL_CLOSED": "Гараар хаасан"}
    wb = Workbook()
    ws = wb.active
    ws.title = site.name[:30]
    ws.append(["Дугаар", "Орсон", "Гарсан", "Хугацаа (мин)", "Дүн (₮)", "НӨАТ (₮)",
               "Хөнгөлөлт (₮)", "Гэрээт", "Төлөв"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for s in rows:
        ws.append([
            s.plate_number,
            (s.entry_time + TZ).strftime("%Y-%m-%d %H:%M"),
            (s.exit_time + TZ).strftime("%Y-%m-%d %H:%M") if s.exit_time else "",
            s.duration_minutes or "",
            float(s.total_fee or 0), float(s.vat_amount or 0), float(s.discount_amount or 0),
            "Тийм" if s.is_registered else "",
            STATUS_MN.get(s.status, s.status),
        ])
    for col, w in zip("ABCDEFGHI", (12, 18, 18, 14, 12, 10, 14, 8, 18)):
        ws.column_dimensions[col].width = w
    return _excel_response(wb, f"sessions_{site.site_code}")


@router.get("/shifts/excel")
def shifts_excel(date_from: str | None = None, date_to: str | None = None,
                 db: Session = Depends(get_db), user: User = Depends(require("reports"))):
    """Касс хаалтын тайлангийн Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    from ..models import CashierShift, Payment
    from sqlalchemy import func as f
    start, end = _range(date_from, date_to)
    shifts = (db.query(CashierShift).filter(CashierShift.opened_at >= start,
                                            CashierShift.opened_at < end)
              .order_by(CashierShift.opened_at.desc()).limit(2000).all())
    wb = Workbook()
    ws = wb.active
    ws.title = "Касс хаалтын тайлан"
    ws.append(["Кассчин", "Төлөв", "Нээсэн цаг", "Хаасан цаг", "Эхэлсэн дүн (₮)",
               "Гүйлгээний тоо", "Бэлэн (₮)", "QPay (₮)", "Карт (₮)", "Нийт орлого (₮)"])
    for c in ws[1]:
        c.font = Font(bold=True)
    grand = 0.0
    for s in shifts:
        totals = dict(db.query(Payment.provider, f.coalesce(f.sum(Payment.amount), 0))
                      .filter(Payment.shift_id == s.id, Payment.status == "PAID")
                      .group_by(Payment.provider).all())
        count = db.query(Payment).filter(Payment.shift_id == s.id, Payment.status == "PAID").count()
        cash, qpay_amt, pos = (float(totals.get(k, 0)) for k in ("CASH", "QPAY", "POS"))
        total = cash + qpay_amt + pos
        grand += total
        ws.append([s.user.username if s.user else "", "Нээлттэй" if s.status == "OPEN" else "Хаагдсан",
                   (s.opened_at + TZ).strftime("%Y-%m-%d %H:%M"),
                   (s.closed_at + TZ).strftime("%Y-%m-%d %H:%M") if s.closed_at else "",
                   float(s.opening_amount or 0), count, cash, qpay_amt, pos, total])
    ws.append(["НИЙТ", "", "", "", "", "", "", "", "", grand])
    ws[f"A{ws.max_row}"].font = Font(bold=True)
    ws[f"J{ws.max_row}"].font = Font(bold=True)
    for col, w in zip("ABCDEFGHIJ", (14, 10, 18, 18, 14, 14, 12, 12, 12, 16)):
        ws.column_dimensions[col].width = w
    return _excel_response(wb, "cashier_shifts")


@router.get("/vat-info")
async def vat_info(user: User = Depends(require("vat", "reports"))):
    """PosAPI getInformation — сугалааны үлдэгдэл, илгээгдээгүй мэдээ (ТЕГ шаардлага №6).
    Frontend үүнийг ашиглан анхааруулга харуулна."""
    from ..config import settings
    from ..services import ebarimt
    info = await ebarimt.get_information()
    warnings = []
    if int(info.get("leftLotteries") or 0) < 500:
        warnings.append(f"Сугалааны дугаар дуусаж байна ({info.get('leftLotteries')} үлдсэн) — "
                        "шинээр авахгүй бол сугалаагүй баримт хэвлэгдэнэ!")
    if int(info.get("unsentCount") or 0) > 0:
        warnings.append(f"Илгээгдээгүй {info.get('unsentCount')} баримт байна — "
                        "3 хоногийн дотор илгээх хуультай.")
    # e-Barimt-ийн 2 суваг: (1) QR/QPay — ebarimt_v3 (бодит, QPay ТЕГ рүү өөрөө илгээнэ),
    # (2) локал PosAPI — картын/бэлэн баримтад (энэ хуудасны сугалаа/мэдээ илгээх хэсэг).
    return {**info, "warnings": warnings,
            "qpay_ebarimt": settings.qpay_ebarimt and not settings.qpay_mock,
            "local_posapi_mock": settings.ebarimt_mock}


@router.post("/vat-send")
async def vat_send(db: Session = Depends(get_db), user: User = Depends(require("vat", "reports"))):
    """Борлуулалтын мэдээг ТЕГ рүү ГАРААР илгээх (ТЕГ шаардлага №5 — гэмтэл саатлын үед)."""
    from ..models import AuditLog
    from ..services import ebarimt
    result = await ebarimt.send_data()
    db.add(AuditLog(username=user.username, action="VAT_SEND_DATA", entity="ebarimt",
                    detail={"result": result.get("message", str(result.get("success")))}))
    db.commit()
    return result


@router.get("/vat-receipts")
def vat_receipts(date_from: str | None = None, date_to: str | None = None,
                 limit: int = 200, db: Session = Depends(get_db),
                 user: User = Depends(require("vat", "reports"))):
    start, end = _range(date_from, date_to)
    rows = (db.query(VatReceipt).filter(VatReceipt.created_at >= start,
                                        VatReceipt.created_at < end)
            .order_by(VatReceipt.created_at.desc()).limit(min(limit, 1000)).all())
    return [to_dict(r) for r in rows]


@router.get("/audit-logs")
def audit_logs(username: str | None = None, action: str | None = None, limit: int = 200,
               db: Session = Depends(get_db), user: User = Depends(require("logs"))):
    q = db.query(AuditLog)
    if username:
        q = q.filter(AuditLog.username == username)
    if action:
        q = q.filter(AuditLog.action == action)
    return [to_dict(a) for a in q.order_by(AuditLog.created_at.desc()).limit(min(limit, 1000)).all()]


@router.get("/audit-logs/excel")
def audit_logs_excel(username: str | None = None, action: str | None = None,
                     db: Session = Depends(get_db), user: User = Depends(require("logs"))):
    """Үйлдлийн логийг Excel болгон татна (ADMIN/FINANCE)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    q = db.query(AuditLog)
    if username:
        q = q.filter(AuditLog.username == username)
    if action:
        q = q.filter(AuditLog.action == action)
    rows = q.order_by(AuditLog.created_at.desc()).limit(10000).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Үйлдлийн лог"
    ws.append(["Огноо", "Хэрэглэгч", "Үйлдэл", "Объект", "Объект ID", "Дэлгэрэнгүй"])
    for c in ws[1]:
        c.font = Font(bold=True)
    import json as _json
    for a in rows:
        ws.append([(a.created_at + TZ).strftime("%Y-%m-%d %H:%M:%S"), a.username, a.action,
                   a.entity or "", a.entity_id or "",
                   _json.dumps(a.detail, ensure_ascii=False) if a.detail else ""])
    for col, w in zip("ABCDEF", (20, 14, 18, 12, 38, 50)):
        ws.column_dimensions[col].width = w
    return _excel_response(wb, "uildliin_log")


@router.get("/lpr-events")
def lpr_events(site_id: str | None = None, limit: int = 200,
               db: Session = Depends(get_db), user: User = Depends(require("logs", "dashboard"))):
    q = db.query(LprEvent)
    if site_id:
        q = q.filter(LprEvent.site_id == site_id)
    return [to_dict(e) for e in q.order_by(LprEvent.created_at.desc()).limit(min(limit, 1000)).all()]
