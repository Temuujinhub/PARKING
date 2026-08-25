"""ТЕГ-ийн мерчант порталын баримтын экспортыг манай баримттай тулгах логик.

Яагаад тусдаа модуль: (1) reports_router аль хэдийн 1,300+ мөр, (2) файл
задлах хэсэг нь оношилгоотой болж томорсон.

ГОЛ ЗАРЧИМ — файлын БАГАНА тогтмол биш. Портал/хувилбар/гар засвараас
хамаарч ДДТД, огноо, дүн хаана байх нь хөвдөг тул урьдын «B=ДДТД, C=огноо,
D=дүн» гэсэн ТОГТМОЛ индекс нэг ч мөр танихгүй болоод «Алдаа гарлаа» гэсэн
ганц мөр буцаадаг байв (хэрэглэгч юу буруу болсныг мэдэх аргагүй). Одоо:
толгой мөрөөр, тэр нь олдохгүй бол өгөгдлийн шинжээр багана танина; танихгүй
бол ЯГ ЮУ уншсанаа (мөрийн жишээ, алгассан шалтгаан бүрийн тоо) буцаана.
"""
import csv as _csv
import io as _io
import re
from collections import Counter
from datetime import date, datetime, time, timedelta

from fastapi import HTTPException

# ДДТД: ТЕГ 33 оронтой дугаарладаг; богиносгосон хувилбарт ч 30-аас багагүй
DDTD_MIN_LEN = 30
_SPACES = re.compile(r"[\s ._/\-]+")


def _norm(v) -> str:
    """Толгойн нэрийг харьцуулах хэлбэрт: жижиг үсэг, зай/цэг/зураас хаяна."""
    return _SPACES.sub("", str(v or "").strip().lower())


def as_ddtd(v) -> str | None:
    """Нүдийг ДДТД болгож уншина. Excel-д 33 оронтой дугаарыг ТОО болгож
    хадгалсан бол 1.52000200900001e+31 болж эвдэрдэг — түүнийг ч барина."""
    if v is None:
        return None
    if isinstance(v, float):
        # Тоо болж хадгалагдсан ДДТД — 15 оронгоос цааш нарийвчлал алдагдсан
        return None if v != int(v) or len(str(int(v))) < DDTD_MIN_LEN else str(int(v))
    if isinstance(v, int) and not isinstance(v, bool):
        s = str(v)
        return s if len(s) >= DDTD_MIN_LEN else None
    s = str(v).strip().replace(" ", "")
    return s if s.isdigit() and len(s) >= DDTD_MIN_LEN else None


_DT_FORMATS = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
               "%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%d-%m-%Y")


def as_dt(v) -> datetime | None:
    """Огноог уншина: datetime нүд, «2026-08-24 06:07:05.0», «2026.08.24»,
    «24/08/2026 06:07», ISO «T» — бүгдийг барина."""
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime.combine(v, time())
    s = str(v or "").strip().replace("T", " ")
    if not s or not s[0].isdigit():
        return None
    head, _, tail = s.partition(" ")
    s = (re.sub(r"[./]", "-", head) + (" " + tail if tail else ""))[:19].strip()
    for f in _DT_FORMATS:
        try:
            return datetime.strptime(s, f)
        except ValueError:
            continue
    return None


def as_num(v) -> float | None:
    """Дүнг уншина: тоо, «8,000.00», «8 000₮», «(1500)» (сөрөг) бүгдийг."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v or "").strip().replace(",", "").replace("₮", "")
    s = _SPACES.sub("", s) if " " in s or " " in s else s
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return None


# ─────────────────────────── Файл унших ────────────────────────────────────

def read_table(filename: str, raw: bytes) -> tuple[list[list], str, str]:
    """Файлыг мөрийн жагсаалт болгож уншина. Буцаана: (мөрүүд, хуудасны нэр, төрөл).

    Форматын алдааг ТОДОРХОЙ хэлнэ — «xlsx уншиж чадсангүй» биш, «энэ бол
    хуучин .xls, Excel дээр нээгээд .xlsx болгож хадгална уу» гэж."""
    name = (filename or "").lower()
    if not raw:
        raise HTTPException(400, "Файл хоосон байна (0 байт) — дахин экспорт хийж үзнэ үү.")
    if raw[:4] == b"\xd0\xcf\x11\xe0":
        raise HTTPException(400, "Энэ бол ХУУЧИН .xls (Excel 97-2003) файл. Excel дээр нээгээд "
                                 "«Save As → Excel Workbook (.xlsx)» гэж хадгалаад дахин оруулна уу.")
    if raw[:5] == b"%PDF-":
        raise HTTPException(400, "PDF файл байна. ТЕГ порталаас Excel (.xlsx) эсвэл CSV "
                                 "экспортыг татаж оруулна уу.")
    if raw[:2] == b"PK":            # zip → xlsx/xlsm
        import openpyxl as _xl
        try:
            wb = _xl.load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as e:      # noqa: BLE001
            raise HTTPException(400, f"Excel файл уншигдсангүй: {str(e)[:150]}. Файл "
                                     "гэмтсэн эсвэл нууц үгтэй байж болзошгүй.")
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        # Идэвхтэй хуудас хоосон бол баримттай хуудсыг өөрөө хайна (олон хуудастай экспорт)
        if len([r for r in rows if any(c is not None for c in r)]) < 2:
            for other in wb.worksheets:
                cand = [list(r) for r in other.iter_rows(values_only=True)]
                if len([r for r in cand if any(c is not None for c in r)]) >= 2:
                    return cand, other.title, "xlsx"
        return rows, ws.title, "xlsx"
    # Текст (CSV/TSV) — ТЕГ портал CSV-ээр ч экспортолдог
    for enc in ("utf-8-sig", "cp1251", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise HTTPException(400, "Файлын төрөл танигдсангүй — .xlsx эсвэл .csv оруулна уу.")
    if "\x00" in text[:2000]:
        raise HTTPException(400, "Файлын төрөл танигдсангүй (хоёртын файл). ТЕГ порталын "
                                 ".xlsx эсвэл .csv экспортыг оруулна уу.")
    sample = text[:4000]
    delim = max(",;\t|", key=lambda d: sample.count(d))
    rows = [r for r in _csv.reader(_io.StringIO(text), delimiter=delim)]
    if not rows:
        raise HTTPException(400, "Файлаас нэг ч мөр уншигдсангүй.")
    return rows, name.rsplit("/", 1)[-1] or "csv", "csv"


# ────────────────────────── Багана таних ───────────────────────────────────

_HDR = {
    "ddtd": ("ддтд", "дтд", "billid", "баримтындугаар", "баримтдугаар", "дугаарбаримт"),
    "dt": ("огноо", "огнооцаг", "date", "createdat", "хугацаа", "гүйлгээнийогноо", "огноохугацаа"),
    "amount": ("нийтдүн", "totalamount", "нийтүнэ", "дүн", "amount", "total", "төлбөр"),
    "vat": ("нөат", "vat", "нөатындүн", "нөатдүн"),
    "src": ("төрөл", "type", "эхсурвалж", "source", "суваг", "постөрөл", "postype", "хэлбэр"),
}
_COL_LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"


def col_name(i: int) -> str:
    """0→A, 25→Z, 26→AA (Excel-ийн баганын нэр — хэрэглэгчид харуулахад)."""
    s = ""
    while True:
        s = _COL_LETTERS[i % 26] + s
        i = i // 26 - 1
        if i < 0:
            return s


def _letter_to_idx(letter: str | None) -> int | None:
    """«C» / «c» / «3» → 0-based индекс. Хэрэглэгчийн гар тохиргоо."""
    if not letter:
        return None
    s = str(letter).strip().upper()
    if s.isdigit():
        return max(0, int(s) - 1)
    idx = 0
    for ch in s:
        if ch not in _COL_LETTERS:
            return None
        idx = idx * 26 + (_COL_LETTERS.index(ch) + 1)
    return idx - 1


def detect_columns(rows: list[list], override: dict | None = None) -> tuple[dict, int]:
    """(багануудын индекс, өгөгдөл эхлэх мөрийн дугаар)-ыг тодорхойлно.

    1) Толгой мөр хайна (эхний 20 мөрөөс танил нэр 2+ таарвал толгой).
    2) Толгойгүй/дутуу бол өгөгдлийн ШИНЖЭЭР танина (30+ оронтой тоо = ДДТД,
       уншигдах огноо = огноо, огнооны БАРУУН талын эхний тоон багана = дүн).
    3) override (хэрэглэгчийн заасан багана) бүхнээс дээгүүр."""
    cols: dict[str, int] = {}
    data_from = 0
    for i, r in enumerate(rows[:20]):
        # Толгой мөрөнд БОДИТ өгөгдөл (ДДТД/огноо) байхгүй — эс бөгөөс «Огноо»
        # гэсэн үг агуулсан өгөгдлийн мөрийг толгой гэж андуурч эхний баримтыг
        # алгасах эрсдэлтэй
        if any(as_ddtd(c) or as_dt(c) for c in r):
            continue
        hit, seen = {}, {_norm(c) for c in r}
        for key, names in _HDR.items():
            for j, c in enumerate(r):
                n = _norm(c)
                if not n:
                    continue
                if key == "amount" and any(v in n for v in _HDR["vat"]):
                    continue          # «НӨАТ-ын дүн» нь НИЙТ дүн биш
                if n in names or any(n.startswith(x) for x in names):
                    hit.setdefault(key, j)
        if len(hit) >= 2 and ("ddtd" in hit or "dt" in hit) and any(seen):
            cols, data_from = hit, i + 1
            break
    body = [r for r in rows[data_from:] if any(c is not None and str(c).strip() for c in r)]
    width = max((len(r) for r in body[:200]), default=0)
    sample = body[:200]
    if "ddtd" not in cols:
        score = [sum(1 for r in sample if as_ddtd(r[j] if j < len(r) else None)) for j in range(width)]
        if score and max(score) > 0:
            cols["ddtd"] = score.index(max(score))
    if "dt" not in cols:
        score = [sum(1 for r in sample if as_dt(r[j] if j < len(r) else None)) for j in range(width)]
        if score and max(score) > 0:
            cols["dt"] = score.index(max(score))
    if "amount" not in cols:
        # Тоон багануудаас: (1) огнооны БАРУУН талынх, (2) тогтмол БИШ утгатайг
        # эрэмбэлж сонгоно. ТТД/салбарын дугаар зэрэг «тогтмол» багана нь бүх
        # мөрд ижил байдаг тул дүн байх магадлал багатай.
        taken = {cols.get("ddtd"), cols.get("dt")}
        ranked = []
        for j in range(width):
            if j in taken:
                continue
            good = [v for v in (as_num(r[j]) if j < len(r) else None for r in sample) if v is not None]
            if len(good) < max(1, len(sample) * 0.6):
                continue
            right = cols.get("dt") is not None and j > cols["dt"]
            ranked.append(((0 if right else 1, 0 if len(set(good)) > 1 else 1, j), j))
        if ranked:
            cols["amount"] = min(ranked)[1]
    for key, letter in (override or {}).items():
        idx = _letter_to_idx(letter)
        if idx is not None:
            cols[key] = idx
    return cols, data_from


def _sample_text(rows: list[list], data_from: int) -> str:
    """Алдааны мессежид оруулах «эхний өгөгдлийн мөр» — юу уншсанаа харуулна."""
    for r in rows[data_from:]:
        if any(c is not None and str(c).strip() for c in r):
            parts = [f"{col_name(j)}=«{str(c)[:34]}»" for j, c in enumerate(r[:14])
                     if c is not None and str(c).strip()]
            return " | ".join(parts)
    return "(өгөгдлийн мөр алга)"


def parse_tax_export(filename: str, raw: bytes, override: dict | None = None) -> tuple[list, dict]:
    """ТЕГ файлыг задлана. Буцаана: (баримтууд, оношилгоо).
    Нэг ч баримт олдоогүй бол ЯГ ЮУ болсныг тайлбарласан 400-г шиднэ."""
    rows, sheet, kind = read_table(filename, raw)
    cols, data_from = detect_columns(rows, override)
    body = [r for r in rows[data_from:] if any(c is not None and str(c).strip() for c in r)]
    tax, skip = [], Counter()
    i_d, i_t, i_a = cols.get("ddtd"), cols.get("dt"), cols.get("amount")
    i_s = cols.get("src")
    lossy = 0
    for r in body:
        cell = lambda j: (r[j] if j is not None and j < len(r) else None)  # noqa: E731
        ddtd, dt, amount = as_ddtd(cell(i_d)), as_dt(cell(i_t)), as_num(cell(i_a))
        if ddtd and isinstance(cell(i_d), (int, float)) and not isinstance(cell(i_d), bool):
            lossy += 1
        if not ddtd:
            skip["ддтд"] += 1
            continue
        if dt is None:
            skip["огноо"] += 1
            continue
        if amount is None:
            skip["дүн"] += 1
            continue
        tax.append({"ddtd": ddtd, "dt": dt, "amount": amount,
                    "src": str(cell(i_s) or "").strip(), "used": False})
    diag = {
        "file": (filename or "")[:80], "kind": kind, "sheet": sheet, "bytes": len(raw),
        "rows": len(body), "parsed": len(tax), "header_row": data_from,
        "columns": {k: col_name(v) for k, v in cols.items() if v is not None},
        "skipped": dict(skip),
        "duplicate_ddtd": len(tax) - len({t["ddtd"] for t in tax}),
        "warnings": [],
    }
    if lossy:
        # Excel float 17 оронтой л нарийвчлалтай — 33 оронтой ДДТД тоо болж
        # хадгалагдвал сүүлийн орнууд 0 болж эвдэрдэг. Тулгалт цаг+дүнгээр
        # явдаг тул ажиллана, гэхдээ экспортын ДДТД баганад найдаж болохгүй.
        diag["warnings"].append(
            f"{lossy} мөрийн ДДТД Excel дээр ТОО болж хадгалагдсан тул сүүлийн орнууд "
            "эвдэрсэн байж болзошгүй (тулгалтад нөлөөлөхгүй — цаг+дүнгээр тулгадаг).")
    if diag["duplicate_ddtd"]:
        diag["warnings"].append(f"Файлд {diag['duplicate_ddtd']} давхардсан ДДТД байна.")
    if not tax:
        raise HTTPException(400, _no_rows_message(diag, rows, data_from, skip))
    return tax, diag


def _no_rows_message(diag: dict, rows: list, data_from: int, skip: Counter) -> str:
    """«Юуны алдаа вэ» гэдгийг мөрийн жишээ + шалтгаанаар нь тайлбарлана."""
    cols = diag["columns"]
    lines = [
        "ТЕГ файлаас нэг ч баримт олдсонгүй.",
        f"Файл: {diag['file'] or '—'} ({diag['kind']}, хуудас «{diag['sheet']}», "
        f"{diag['rows']} өгөгдлийн мөр)",
        f"Танигдсан багана: ДДТД={cols.get('ddtd', 'ОЛДСОНГҮЙ')} · "
        f"Огноо={cols.get('dt', 'ОЛДСОНГҮЙ')} · Дүн={cols.get('amount', 'ОЛДСОНГҮЙ')}",
    ]
    if skip:
        lines.append("Алгассан мөр: " + ", ".join(f"{k} уншигдаагүй {v}" for k, v in skip.items()))
    lines.append("Эхний мөр: " + _sample_text(rows, data_from))
    if skip.get("ддтд"):
        lines.append("💡 ДДТД (33 оронтой) багана Excel дээр ТОО болж хадгалагдвал "
                     "1.52E+31 болж эвдэрдэг — тэр баганыг «Text» болгож дахин экспортлох, "
                     "эсвэл доорх «Багана гараар заах» хэсгээс зөв баганыг сонгоно уу.")
    elif skip.get("огноо") and "dt" in cols:
        lines.append("💡 Огнооны багана танигдсан ч уншигдахгүй байна "
                     "(дэмждэг форматууд: 2026-08-24 06:07:05, 2026.08.24 06:07, "
                     "24/08/2026 06:07).")
    else:
        lines.append("💡 Багана өөр байрлалтай бол «Багана гараар заах» хэсэгт "
                     "ДДТД/Огноо/Дүнгийн баганын үсгийг (ж: B, C, D) бичиж дахин оролдоно уу.")
    return "\n".join(lines)


# ──────────────────────────── Тулгалт ──────────────────────────────────────

def match_pass(tax: list, ours: list, shift: float, tol: int) -> dict:
    """Нэг цагийн шилжилтээр тулгана.

    ДДТД-ЭЭР ТУЛГАДАГГҮЙ — суваг бүр (QPay, msgbill) операторын кодтой билл
    буцаадаг бол ТЕГ такспэерийн ТТД + өөрийн counter-оор ӨӨР ДДТД олгодог.
    Гэхдээ таарсан хос бүрийн хоёр ДДТД-г ХАРЬЦУУЛЖ, сувгаар нь тоолно —
    «манай дугаар ТЕГ-ийнхтэй таарч байна уу» гэдгийг хэмжих цорын ганц зам."""
    for t in tax:
        t["used"] = False
        t.pop("ours", None)
    for t in tax:
        t.setdefault("shift", shift)
    matched, ddtd_equal, un_ours, cancelled = 0, 0, [], 0
    by_provider: dict[str, dict] = {}

    def when(t):
        """Мөр бүрийн цаг өөрийн КАССЫН шилжилтээр (файл дотор холилдож ирдэг)."""
        return t["dt"] + timedelta(hours=t["shift"])

    for rec, pay, plate, site_name in ours:
        if pay.paid_at is None:
            continue
        if rec.status == "CANCELLED":
            cancelled += 1          # ТЕГ-т цуцлагдсан тул файлд байхгүй нь ХЭВИЙН
            continue
        cand = [t for t in tax if not t["used"] and abs(float(rec.amount) - t["amount"]) < 1
                and abs((when(t) - pay.paid_at).total_seconds()) <= tol]
        if cand:
            best = min(cand, key=lambda x: abs((when(x) - pay.paid_at).total_seconds()))
            best["used"] = True
            best["ours"] = (rec, pay, plate, site_name)
            matched += 1
            same = rec.ebarimt_id == best["ddtd"]
            ddtd_equal += int(same)
            prov = rec.provider or "POSAPI"
            st = by_provider.setdefault(prov, {"matched": 0, "equal": 0, "sample": None})
            st["matched"] += 1
            st["equal"] += int(same)
            if st["sample"] is None and not same:
                # Ялгаатай хосын ЖИШЭЭ — хоёр дугаарыг зэрэгцүүлж харуулна
                st["sample"] = {"ours": rec.ebarimt_id or "", "tax": best["ddtd"]}
        else:
            un_ours.append({"paid_at": pay.paid_at.isoformat(), "plate": plate,
                            "site_name": site_name, "amount": float(rec.amount),
                            "status": rec.status, "provider": rec.provider or "POSAPI",
                            "ebarimt_id": rec.ebarimt_id})
    return {"matched": matched, "ddtd_equal": ddtd_equal, "unmatched_ours": un_ours,
            "cancelled": cancelled, "by_provider": by_provider}


def clip_to_file(tax: list, ours: list, shift: float, tol: int) -> tuple[list, int]:
    """Манай баримтуудыг ФАЙЛЫН хамрах хугацаанд тайрна.

    Асуулга нь цагийн шилжилт (0 / −TZ) аль нь ч таарахаар ±9 цагийн НӨӨЦТЭЙ
    татдаг. Тэр нөөц дэх баримтуудыг «манайд бий, ТЕГ-д алга» гэж тоолох нь
    ХУДАЛ — файл тэр хугацааг огт хамраагүй байхад «зөрүү» болж харагдана.
    (2026-08-25: 121 «зөрүү»-гийн ихэнх нь яг энэ нөөцийн баримтууд байв.)"""
    times = [t["dt"] + timedelta(hours=t.get("shift", shift)) for t in tax]
    lo, hi = min(times) - timedelta(seconds=tol), max(times) + timedelta(seconds=tol)
    inside = [o for o in ours if o[1].paid_at is not None and lo <= o[1].paid_at <= hi]
    return inside, len(ours) - len(inside)


def best_shift(tax: list, ours: list, candidates: list[float], tol: int) -> dict:
    """Файлын цаг UTC уу, УБ-ын локал уу гэдгийг ТААРАЛТААР нь өөрөө сонгоно.
    (Урьд нь tz_shift=0 тогтмол байсан тул локал цагтай экспорт 0 таарч,
    «бүгд зөрж байна» гэж харагддаг байв.)

    Шилжилтийг сонгосны ДАРАА манай талыг файлын хамрах хугацаанд тайрч,
    эцсийн тулгалтыг тэр багц дээр хийнэ."""
    # ── Цагийн шилжилтийг КАСС (POS) тус бүрээр тодорхойлно ──────────────────
    # 2026-08-25: ТЕГ-ийн нэг экспорт дотор ЦАГИЙН БҮС ХОЛИЛДОЖ ирдэг нь
    # батлагдав — QPay-ийн касс UTC-ээр, msgbill-ийн касс УБ локал цагаар
    # (+8ц) бичигдсэн. Нэг ерөнхий шилжилт сонговол нөгөө кассын БҮХ мөр
    # «манайд алга» болж 76 хуурамч зөрүү гардаг байв.
    tail = pos_tail_len([t["ddtd"] for t in tax])
    groups: dict[str, list] = {}
    for t in tax:
        groups.setdefault(t["ddtd"][-tail:] if tail else "*", []).append(t)
    for key, rows in groups.items():
        best_sh, best_n = candidates[0], -1
        for sh in candidates:
            for x in rows:          # ЗӨВХӨН энэ бүлгийнх — өмнөх бүлгийн
                x["shift"] = sh     # шилжилтийг дарж бичихгүй
            n = match_pass(rows, ours, sh, tol)["matched"]
            if n > best_n:
                best_sh, best_n = sh, n
        for x in rows:
            x["group"] = key
            x["shift"] = best_sh
    shift = min((t["shift"] for t in tax), key=lambda v: abs(v)) if tax else candidates[0]
    inside, outside = clip_to_file(tax, ours, shift, tol)
    res = match_pass(tax, inside, shift, tol)     # сонгосон шилжилтийг дахин тавина
    return {**res, "shift": shift, "ours_in_window": len(inside), "outside_window": outside,
            "inside": inside,
            "group_shifts": {k: rows[0]["shift"] for k, rows in groups.items()}}


def pos_tail_len(ids: list[str]) -> int | None:
    """ДДТД-ийн төгсгөл дэх КАССЫН дугаарын уртыг өгөгдлөөс олно.

    Шалгуур: бүлэг цөөн (≤6) БӨГӨӨД бүлэг бүрд дунджаар 4+ мөр ноогдоно.
    Эс бөгөөс дарааллын дугаарыг «касс» гэж андуурч мөр бүрийг тусад нь
    бүлэглэнэ. Тохирох урт олдоогүй бол None — бүх мөр НЭГ бүлэг."""
    ids = [i for i in ids if i]
    if len(ids) < 8:
        return None
    found = None
    for ln in range(6, 13):
        n = len({i[-ln:] for i in ids})
        if n <= 6 and n * 4 <= len(ids):
            found = ln
    return found


def pos_groups(ids: list[str], tail: int | None = None) -> dict:
    """ДДТД-үүдийг кассын дугаараар бүлэглэнэ (харуулах зориулалттай).

    tail: бүх файлаас тооцсон урт. Дэд багц (ж: таараагүй мөрүүд) дээр дахин
    тооцвол өөр урт гарч, харьцуулах боломжгүй болно."""
    ids = [i for i in ids if i]
    if not ids:
        return {}
    ln = tail or pos_tail_len(ids)
    return dict(Counter(i[-ln:] for i in ids)) if ln else {"(нэг бүлэг)": len(ids)}


# «ТЕГ-д бий, манайд алга» мөрийн ШАЛТГААНЫГ манай ТӨЛБӨРИЙН бүртгэлээр тайлах.
# Баримтын бүртгэл байхгүй ч ТӨЛБӨР нь байвал — баримт үүссэн боловч манайд
# хадгалагдаагүй (эсвэл давхар үүссэн) гэсэн үг. Энэ ялгааг гараар хөөх нь
# хэдэн цагийн ажил байсан.
VERDICTS = {
    "DUPLICATE": "ДАВХАР БАРИМТ — тухайн төлбөрт манайд баримт БИЙ (өөр ДДТД-тэй)",
    "PAYMENT_NO_RECEIPT": "БАРИМТГҮЙ ТӨЛБӨР — төлбөр авсан ч баримтын бүртгэл алга",
    "RECEIPT_NOT_MATCHED": "Баримт бий ч цаг/дүнгээр таарсангүй",
    "OTHER_SCOPE": "ӨӨР ЗОГСООЛ — сонгосон түрээслэгчийн шүүлтээс гадуур",
    "NO_PAYMENT": "Манайд төлбөр ч алга — өөр систем/POS",
}


def explain_unmatched_tax(left: list, probe: list, shift: float, tol: int,
                          matched_pay_ids: set, scope_ids: set | None) -> list:
    """Тохироогүй ТЕГ мөр бүрийг манай ТӨЛБӨРүүдтэй (баримттай эсэхээс үл хамааран)
    цаг+дүнгээр тулгаж ангилна. probe: {id, paid_at, amount, site_id, ...} жагсаалт."""
    off = timedelta(hours=shift)
    out = []
    for t in left:
        when = t["dt"] + off
        cand = [p for p in probe if abs(p["amount"] - t["amount"]) < 1
                and abs((p["paid_at"] - when).total_seconds()) <= tol]
        row = {"dt": t["dt"].isoformat(), "amount": t["amount"], "src": t["src"],
               "ddtd": t["ddtd"], "verdict": "NO_PAYMENT", "plate": "", "site_name": "",
               "method": "", "our_ddtd": "", "payment_id": None}
        if cand:
            # Аль хэдийн таарсан төлбөрийг СҮҮЛД нь авна — эс бөгөөс завгүй цагт
            # ижил дүнтэй хоёр төлбөр байхад хуурамч «давхардал» гарна
            p = min(cand, key=lambda x: (x["id"] in matched_pay_ids,
                                         abs((x["paid_at"] - when).total_seconds())))
            if scope_ids is not None and p["site_id"] not in scope_ids:
                v = "OTHER_SCOPE"
            elif not p["receipts"]:
                v = "PAYMENT_NO_RECEIPT"
            elif p["id"] in matched_pay_ids:
                v = "DUPLICATE"
            else:
                v = "RECEIPT_NOT_MATCHED"
            row.update({"verdict": v, "plate": p["plate"] or "", "site_name": p["site_name"] or "",
                        "method": f'{p["provider"]}/{p["method"]}',
                        "our_ddtd": ", ".join(x for x in p["receipts"] if x)[:80],
                        # Бөөнөөр цуцлахад хэрэгтэй — UI энэ ID-аар хүсэлт илгээнэ
                        "payment_id": p["id"]})
        out.append(row)
    return out


def reconcile_excel(tax: list, ours: list, un_ours: list, tol: int, tz: int,
                    title: str = "", shift: float = 0, explained: dict | None = None):
    """Тулгалтын нэгтгэсэн xlsx: (1) ТЕГ+манай мөр зэрэгцээ, (2) манайд бий/ТЕГ-д алга,
    (3) дүгнэлт. Огноог ЛОКАЛ (УБ) цагаар харуулна.

    shift: тулгалтад ашигласан цагийн шилжилт. ТЕГ-ийн огноог эхлээд түүгээр
    UTC болгож, дараа нь локал болгоно — эс бөгөөс локал цагтай файл ХОЁР УДАА
    шилжиж (+16ц) харагдана."""
    from urllib.parse import quote
    import openpyxl as _xl
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Font, PatternFill

    wb = _xl.Workbook()
    bold = Font(bold=True)
    warn = PatternFill("solid", fgColor="FFF2CC")
    loc = timedelta(hours=tz)
    tax_loc = timedelta(hours=shift) + loc     # файлын цаг → UTC → локал

    ws = wb.active
    ws.title = "Тулгалт"
    ws.append(["ТЕГ огноо (УБ цаг)", "Дүн ₮", "ТЕГ ДДТД", "ТЕГ эх сурвалж",
               "Тулгалт", "Машины дугаар", "Зогсоол", "Манай суваг", "Манай ДДТД",
               "Сугалаа", "Манай төлөв", "Төлсөн (УБ цаг)", "Шалтгаан (таараагүй бол)"])
    for c in ws[1]:
        c.font = bold
    for t in sorted(tax, key=lambda x: x["dt"]):
        o = t.get("ours")
        row = [(t["dt"] + tax_loc).strftime("%Y-%m-%d %H:%M:%S"), t["amount"], t["ddtd"], t["src"]]
        if o:
            rec, pay, plate, site_name = o
            row += ["ТААРСАН", plate or "", site_name or "", rec.provider or "POSAPI",
                    rec.ebarimt_id or "", rec.lottery_code or "", rec.status,
                    (pay.paid_at + loc).strftime("%Y-%m-%d %H:%M:%S") if pay.paid_at else ""]
        else:
            e = explained.get(t["ddtd"]) if explained else None
            row += ["МАНАЙД АЛГА", (e or {}).get("plate", ""), (e or {}).get("site_name", ""),
                    (e or {}).get("method", ""), (e or {}).get("our_ddtd", ""), "", "", "",
                    VERDICTS.get((e or {}).get("verdict", ""), "")]
        ws.append(row)
        if not o:
            for c in ws[ws.max_row]:
                c.fill = warn

    ws2 = wb.create_sheet("Манайд бий - ТЕГ-д алга")
    ws2.append(["Төлсөн (УБ цаг)", "Машины дугаар", "Зогсоол", "Дүн ₮", "Суваг",
                "Манай төлөв", "Манай ДДТД", "Тайлбар"])
    for c in ws2[1]:
        c.font = bold
    for r in sorted(un_ours, key=lambda x: x["paid_at"]):
        note = ("MOCK/амжилтгүй баримт" if r["status"] != "SENT" else
                "Өөр ТТД дээр бүртгэгдсэн байж болно" if r["provider"] == "QPAY" else "")
        dt = datetime.fromisoformat(r["paid_at"]) + loc
        ws2.append([dt.strftime("%Y-%m-%d %H:%M:%S"), r["plate"] or "", r.get("site_name") or "",
                    r["amount"], r["provider"], r["status"], r["ebarimt_id"] or "", note])

    ws3 = wb.create_sheet("Дүгнэлт")
    matched = sum(1 for t in tax if t.get("ours"))
    lines = [
        ("Тулгалт", title or "—"),
        ("ТЕГ файлын баримт", len(tax)),
        ("Манай баримт (тухайн хугацаанд)", len(ours)),
        ("Таарсан (цаг ±%dс + дүн)" % tol, matched),
        ("ТЕГ-д бий, манайд алга", len(tax) - matched),
        ("Манайд бий, ТЕГ-д алга", len(un_ours)),
        ("", ""),
        ("ТЕГ эх сурвалжаар:", ""),
        *[(f"  {k or '(хоосон)'}", v) for k, v in Counter(t["src"] for t in tax).items()],
        ("", ""),
        ("ТАЙЛБАР: ДДТД хоорондоо таарахгүй нь ХЭВИЙН — QPay/msgbill операторын кодтой", ""),
        ("билл буцаадаг бол ТЕГ такспэерийн ТТД + өөрийн дугаарлалтаар бүртгэдэг.", ""),
        ("Тулгалтыг цаг + дүнгээр хийсэн.", ""),
    ]
    for a, b in lines:
        ws3.append([a, b])
    ws3["A1"].font = bold
    for wsx in (ws, ws2):
        for col, w in zip("ABCDEFGHIJKLM", (20, 10, 36, 16, 13, 14, 16, 11, 36, 12, 10, 20, 46)):
            wsx.column_dimensions[col].width = w
    ws3.column_dimensions["A"].width = 70

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    day = (min(t["dt"] for t in tax) + tax_loc).strftime("%Y%m%d") if tax else "x"
    fname = f"ebarimt-tulgalt-{day}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}; filename*=UTF-8''{quote(fname)}"})
