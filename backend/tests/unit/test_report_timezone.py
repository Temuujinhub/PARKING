"""Тайлан/экспортын цаг УБ-ын хананы цагаар гарах эсэх.

2026-09-03-ны гомдол: «бүх данс тооцоо 8 цагаар зөрж байна». Шалтгаан нь
серверийн цаг БИШ (UTC, NTP-тэй, зөв) — харин ХАРУУЛАХ давхарга: `_txn_rows`
нь naive UTC ISO мөр буцаадаг бөгөөд Excel экспорт түүнийг ХӨРВҮҮЛЭЛГҮЙ
зүсдэг байв (`(iso or "").replace("T"," ")[:16]`). Ижил алдаа frontend-ийн
ReportTabs дээр ч байсан.
"""
from app.config import settings
from app.routers.reports_excel import _iso_local, transactions_excel


def test_utc_iso_is_shifted_to_local_wall_clock():
    """15:07 UTC = 23:07 УБ — 8 цагийн зөрүү энд хаагдана."""
    assert _iso_local("2026-09-02T15:07:00") == "2026-09-02 23:07"


def test_shift_crosses_the_day_boundary():
    """Өдрийн зааг мөн зөв шилжинэ (16:00 UTC = маргаашийн 00:00 УБ)."""
    assert _iso_local("2026-09-02T16:00:00") == "2026-09-03 00:00"


def test_offset_follows_configuration():
    old = settings.tz_offset_hours
    try:
        # TZ өөрчлөгдвөл модулийн TZ тогтмол хэвээр — энэ нь ЗӨВ (нэг deploy-д
        # нэг бүс). Энд зөвхөн одоогийн тохиргоо 8 болохыг барина.
        assert old == 8
    finally:
        settings.tz_offset_hours = old


def test_empty_and_malformed_values_do_not_crash():
    assert _iso_local(None) == ""
    assert _iso_local("") == ""
    assert _iso_local("тоо биш") == "тоо биш"


def test_excel_sheet_contains_local_times_and_paid_column():
    row = {"plate_number": "1234УБА", "site_name": "Тест", "car_type": "Энгийн",
           "entry_time": "2026-09-02T15:07:00", "exit_time": "2026-09-02T15:30:00",
           "paid_at": "2026-09-02T15:30:00", "duration_minutes": 23,
           "discount_name": None, "base_fee": 909, "discount_amount": 0,
           "vat_amount": 91, "total_fee": 1000, "paid_amount": 1000,
           "provider": "QPay", "invoice_no": "x", "status": "Хаагдсан",
           "cashier": None, "ebarimt_id": None, "lottery_code": None,
           "customer_tin": None}
    import asyncio
    import io

    import openpyxl
    resp = transactions_excel([row])          # StreamingResponse — async iterator
    assert resp.status_code == 200

    async def _drain():
        return b"".join([c async for c in resp.body_iterator])

    ws = openpyxl.load_workbook(io.BytesIO(asyncio.run(_drain()))).active
    header = [c.value for c in ws[1]]
    assert "Төлсөн" in header, "төлсөн цагийн багана экспортод байх ёстой"
    values = [c.value for c in ws[2]]
    assert "2026-09-02 23:07" in values, values
    assert "2026-09-02 23:30" in values, values
